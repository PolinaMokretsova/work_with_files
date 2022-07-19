import os
import zipfile
import csv
from PyPDF2 import PdfReader
from openpyxl import load_workbook


def test_add_zip_files():
    my_zip = zipfile.ZipFile('resourses/my_zip.zip', 'w')
    for folder, subfolders, files in os.walk('resourses'):

        for file in files:
            if file.endswith('.pdf') or file.endswith('.xlsx') or file.endswith('.csv'):
                my_zip.write(os.path.join(folder, file), os.path.relpath
                (os.path.join(folder, file), 'resources'), compress_type = zipfile.ZIP_DEFLATED)
    my_zip.close()


def test_unzip_files():
    fantasy_zip = zipfile.ZipFile('resourses/my_zip.zip')
    fantasy_zip.extractall('unzipped/')
    fantasy_zip.close()


def test_read_pdffile():
    reader = PdfReader('unzipped/resourses/docs-pytest-org-en-latest.pdf')
    page = reader.pages[0]
    text = page.extract_text()
    assert '2022' in text


def test_read_xlsxfile():
    workbook = load_workbook('unzipped/resourses/file_example_XLSX_10.xlsx')
    sheet = workbook.active
    name = sheet.cell(row=3, column=3).value
    assert 'Hashimoto' == name


def test_read_csvfile():
    with open('unzipped/resourses/username.csv') as f:
        reader = csv.reader(f)
        headers = next(reader)
    assert 'Username' in str(headers)